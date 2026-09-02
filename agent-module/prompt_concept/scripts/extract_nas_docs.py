"""Extract CQR knowledge into prompt/data/. Local-first; NAS is optional fallback."""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[3]
TOOLS = str(REPO / "tools")
if TOOLS not in sys.path:
    sys.path.insert(0, TOOLS)
from operator_config import nas_dev_xlsx, nas_po_dir  # noqa: E402

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

CONCEPT_ROOT = Path(__file__).resolve().parent.parent
OUT = CONCEPT_ROOT.parent / "data"
LOCAL_SOURCE = OUT / "source"


def first_existing(candidates: list[Path]) -> Path | None:
    for path in candidates:
        if path.exists():
            return path
    return None


def resolve_dev_xlsx() -> Path | None:
    configured = nas_dev_xlsx(REPO)
    candidates = []
    if configured:
        candidates.append(configured)
    candidates.append(LOCAL_SOURCE / "cqr_development_direction.xlsx")
    if LOCAL_SOURCE.is_dir():
        candidates.extend(sorted(LOCAL_SOURCE.glob("*.xlsx")))
    return first_existing(candidates)


def resolve_po_dir() -> Path | None:
    configured = nas_po_dir(REPO)
    candidates = []
    if configured:
        candidates.append(configured)
    candidates.extend([LOCAL_SOURCE / "2_CLOTHING", LOCAL_SOURCE / "po_clothing"])
    return first_existing(candidates)


def dump_workbook(
    path: Path, out_name: str, max_rows: int = 40, max_sheets: int = 8
) -> None:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    lines = [f"# Source: {path.name}", f"Sheets: {', '.join(wb.sheetnames)}", ""]
    for sn in wb.sheetnames[:max_sheets]:
        ws = wb[sn]
        lines.append(f"## {sn}")
        n = 0
        for row in ws.iter_rows(values_only=True):
            vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if vals:
                lines.append(" | ".join(vals[:14]))
                n += 1
                if n >= max_rows:
                    break
        lines.append("")
    wb.close()
    (OUT / out_name).write_text("\n".join(lines), encoding="utf-8")


def scan_po_dir(po_dir: Path) -> None:
    lines = ["# Clothing PO scan (local extract)", f"Source folder: {po_dir.name}", ""]
    files = sorted(
        po_dir.rglob("*"),
        key=lambda p: p.stat().st_mtime if p.is_file() else 0,
        reverse=True,
    )
    listed = 0
    for p in files:
        if not p.is_file():
            continue
        if p.suffix.lower() not in {".xlsx", ".xls", ".xlsm", ".pdf"}:
            continue
        rel = p.relative_to(po_dir)
        lines.append(f"- {rel} | {p.stat().st_size} bytes")
        listed += 1
        if listed >= 80:
            break
    (OUT / "po_clothing_index.txt").write_text("\n".join(lines), encoding="utf-8")

    samples = [
        p
        for p in files
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xls", ".xlsm"}
    ][:5]
    for i, sample in enumerate(samples, 1):
        try:
            dump_workbook(sample, f"po_sample_{i}.txt", max_rows=30, max_sheets=3)
        except OSError as exc:
            (OUT / f"po_sample_{i}.txt").write_text(
                f"ERROR reading {sample.name}: {exc}", encoding="utf-8"
            )


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    LOCAL_SOURCE.mkdir(parents=True, exist_ok=True)

    dev = resolve_dev_xlsx()
    if dev:
        dump_workbook(dev, "cqr_development_direction.txt", max_rows=50, max_sheets=10)
        print("Development direction:", dev)
    else:
        msg = (
            "MISSING development xlsx.\n"
            "Set CQR_DEV_XLSX, copy the workbook to data/source/, or fill _local/operator.json nas.dev_xlsx."
        )
        (OUT / "cqr_development_direction.txt").write_text(msg, encoding="utf-8")
        print(msg)

    po = resolve_po_dir()
    if po:
        scan_po_dir(po)
        print("PO scan:", po)
    else:
        (OUT / "po_clothing_index.txt").write_text(
            "# Clothing PO scan\n\nNo local PO folder found. Optional: copy 2_CLOTHING into data/source/po_clothing/",
            encoding="utf-8",
        )
        print("PO folder not found (optional)")

    print("Wrote extracts to", OUT)


if __name__ == "__main__":
    main()
